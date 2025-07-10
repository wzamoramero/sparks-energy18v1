# -*- coding: utf-8 -*-

import base64
import io
from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError

try:
    import xlrd
    import openpyxl
except ImportError:
    xlrd = None
    openpyxl = None


class ImportConsumptionWizard(models.TransientModel):
    _name = 'sparks.import.consumption.wizard'
    _description = 'Import Monthly Consumption Data Wizard'

    quotation_id = fields.Many2one(
        'sparks.solar.quotation',
        string='Solar Quotation',
        required=True
    )
    
    import_file = fields.Binary(
        string='Import File',
        required=True,
        help="Excel file (.xlsx or .xls) with monthly consumption data"
    )
    
    filename = fields.Char(string='Filename')
    
    file_type = fields.Selection([
        ('excel', 'Excel File (.xlsx/.xls)'),
        ('csv', 'CSV File (.csv)')
    ], string='File Type', default='excel')
    
    has_header = fields.Boolean(
        string='File has header row',
        default=True,
        help="Check if the first row contains column headers"
    )
    
    month_column = fields.Integer(
        string='Month Column',
        default=1,
        help="Column number for month data (1-based)"
    )
    
    consumption_column = fields.Integer(
        string='Consumption Column',
        default=2,
        help="Column number for consumption data (1-based)"
    )
    
    cost_column = fields.Integer(
        string='Cost Column',
        default=3,
        help="Column number for cost data (1-based, optional)"
    )
    
    preview_data = fields.Text(
        string='Preview',
        readonly=True,
        help="Preview of the data to be imported"
    )
    
    state = fields.Selection([
        ('upload', 'Upload File'),
        ('preview', 'Preview Data'),
        ('done', 'Import Complete')
    ], default='upload')

    @api.onchange('import_file', 'filename')
    def _onchange_import_file(self):
        if self.import_file and self.filename:
            # Auto-detect file type
            if self.filename.lower().endswith(('.xlsx', '.xls')):
                self.file_type = 'excel'
            elif self.filename.lower().endswith('.csv'):
                self.file_type = 'csv'
            
            # Generate preview
            self._generate_preview()

    def _generate_preview(self):
        """Generate a preview of the imported data"""
        if not self.import_file:
            return
        
        try:
            file_data = base64.b64decode(self.import_file)
            
            if self.file_type == 'excel':
                preview = self._preview_excel_data(file_data)
            else:
                preview = self._preview_csv_data(file_data)
            
            self.preview_data = preview
            
        except Exception as e:
            self.preview_data = f"Error reading file: {str(e)}"

    def _preview_excel_data(self, file_data):
        """Preview Excel file data"""
        if not openpyxl:
            raise UserError(_("Please install openpyxl library to import Excel files"))
        
        try:
            # Try openpyxl first (for .xlsx)
            workbook = openpyxl.load_workbook(io.BytesIO(file_data))
            sheet = workbook.active
            
            preview_lines = []
            max_rows = 10  # Limit preview to 10 rows
            
            for row_idx, row in enumerate(sheet.iter_rows(max_row=max_rows, values_only=True), 1):
                if row_idx == 1 and self.has_header:
                    preview_lines.append(f"Header: {' | '.join(str(cell) for cell in row if cell is not None)}")
                else:
                    preview_lines.append(f"Row {row_idx}: {' | '.join(str(cell) for cell in row if cell is not None)}")
            
            return '\n'.join(preview_lines)
            
        except Exception:
            # Fallback to xlrd for .xls files
            if not xlrd:
                raise UserError(_("Please install xlrd library to import .xls files"))
            
            workbook = xlrd.open_workbook(file_contents=file_data)
            sheet = workbook.sheet_by_index(0)
            
            preview_lines = []
            max_rows = min(10, sheet.nrows)
            
            for row_idx in range(max_rows):
                row_data = [str(sheet.cell_value(row_idx, col_idx)) for col_idx in range(sheet.ncols)]
                if row_idx == 0 and self.has_header:
                    preview_lines.append(f"Header: {' | '.join(row_data)}")
                else:
                    preview_lines.append(f"Row {row_idx + 1}: {' | '.join(row_data)}")
            
            return '\n'.join(preview_lines)

    def _preview_csv_data(self, file_data):
        """Preview CSV file data"""
        import csv
        
        # Decode file data
        text_data = file_data.decode('utf-8-sig')
        csv_reader = csv.reader(io.StringIO(text_data))
        
        preview_lines = []
        max_rows = 10
        
        for row_idx, row in enumerate(csv_reader):
            if row_idx >= max_rows:
                break
            
            if row_idx == 0 and self.has_header:
                preview_lines.append(f"Header: {' | '.join(row)}")
            else:
                preview_lines.append(f"Row {row_idx + 1}: {' | '.join(row)}")
        
        return '\n'.join(preview_lines)

    def action_preview(self):
        """Show preview of data to be imported"""
        self._generate_preview()
        self.state = 'preview'
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': self.env.context,
        }

    def action_import(self):
        """Import the consumption data"""
        if not self.import_file:
            raise UserError(_("Please select a file to import"))
        
        # Clear existing consumption lines
        self.quotation_id.consumption_line_ids.unlink()
        
        try:
            file_data = base64.b64decode(self.import_file)
            
            if self.file_type == 'excel':
                consumption_data = self._import_excel_data(file_data)
            else:
                consumption_data = self._import_csv_data(file_data)
            
            # Create consumption lines
            self._create_consumption_lines(consumption_data)
            
            self.state = 'done'
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Import Successful'),
                    'message': _('%d consumption records imported successfully') % len(consumption_data),
                    'type': 'success',
                }
            }
            
        except Exception as e:
            raise UserError(_("Error importing file: %s") % str(e))

    def _import_excel_data(self, file_data):
        """Import data from Excel file"""
        consumption_data = []
        
        try:
            # Try openpyxl first
            workbook = openpyxl.load_workbook(io.BytesIO(file_data))
            sheet = workbook.active
            
            start_row = 2 if self.has_header else 1
            
            for row in sheet.iter_rows(min_row=start_row, values_only=True):
                if not any(row):  # Skip empty rows
                    continue
                
                try:
                    month = self._parse_month(row[self.month_column - 1])
                    consumption = float(row[self.consumption_column - 1] or 0)
                    cost = float(row[self.cost_column - 1] or 0) if len(row) >= self.cost_column else 0
                    
                    consumption_data.append({
                        'month': month,
                        'energy_kwh': consumption,
                        'energy_cost': cost
                    })
                    
                except (ValueError, IndexError, TypeError):
                    continue  # Skip invalid rows
                    
        except Exception:
            # Fallback to xlrd
            if not xlrd:
                raise UserError(_("Please install xlrd library to import .xls files"))
            
            workbook = xlrd.open_workbook(file_contents=file_data)
            sheet = workbook.sheet_by_index(0)
            
            start_row = 1 if self.has_header else 0
            
            for row_idx in range(start_row, sheet.nrows):
                try:
                    month = self._parse_month(sheet.cell_value(row_idx, self.month_column - 1))
                    consumption = float(sheet.cell_value(row_idx, self.consumption_column - 1) or 0)
                    cost = float(sheet.cell_value(row_idx, self.cost_column - 1) or 0) if sheet.ncols >= self.cost_column else 0
                    
                    consumption_data.append({
                        'month': month,
                        'energy_kwh': consumption,
                        'energy_cost': cost
                    })
                    
                except (ValueError, IndexError, TypeError):
                    continue  # Skip invalid rows
        
        return consumption_data

    def _import_csv_data(self, file_data):
        """Import data from CSV file"""
        import csv
        consumption_data = []
        
        text_data = file_data.decode('utf-8-sig')
        csv_reader = csv.reader(io.StringIO(text_data))
        
        # Skip header if present
        if self.has_header:
            next(csv_reader, None)
        
        for row in csv_reader:
            if not any(row):  # Skip empty rows
                continue
            
            try:
                month = self._parse_month(row[self.month_column - 1])
                consumption = float(row[self.consumption_column - 1] or 0)
                cost = float(row[self.cost_column - 1] or 0) if len(row) >= self.cost_column else 0
                
                consumption_data.append({
                    'month': month,
                    'energy_kwh': consumption,
                    'energy_cost': cost
                })
                
            except (ValueError, IndexError, TypeError):
                continue  # Skip invalid rows
        
        return consumption_data

    def _parse_month(self, month_value):
        """Parse month value to standard format"""
        if isinstance(month_value, (int, float)):
            month_num = int(month_value)
            if 1 <= month_num <= 12:
                return str(month_num)
        
        if isinstance(month_value, str):
            month_value = month_value.strip().lower()
            
            # Try to match month names
            month_names = {
                'january': '1', 'jan': '1', 'enero': '1',
                'february': '2', 'feb': '2', 'febrero': '2',
                'march': '3', 'mar': '3', 'marzo': '3',
                'april': '4', 'apr': '4', 'abril': '4',
                'may': '5', 'mayo': '5',
                'june': '6', 'jun': '6', 'junio': '6',
                'july': '7', 'jul': '7', 'julio': '7',
                'august': '8', 'aug': '8', 'agosto': '8',
                'september': '9', 'sep': '9', 'septiembre': '9',
                'october': '10', 'oct': '10', 'octubre': '10',
                'november': '11', 'nov': '11', 'noviembre': '11',
                'december': '12', 'dec': '12', 'diciembre': '12'
            }
            
            for name, num in month_names.items():
                if name in month_value:
                    return num
            
            # Try to parse as number
            try:
                month_num = int(month_value)
                if 1 <= month_num <= 12:
                    return str(month_num)
            except ValueError:
                pass
        
        raise ValueError(f"Invalid month value: {month_value}")

    def _create_consumption_lines(self, consumption_data):
        """Create consumption lines from imported data"""
        for data in consumption_data:
            self.env['sparks.consumption.line'].create({
                'quotation_id': self.quotation_id.id,
                'month': data['month'],
                'energy_kwh': data['energy_kwh'],
                'energy_cost': data['energy_cost']
            })

    def action_back_to_upload(self):
        """Go back to upload step"""
        self.state = 'upload'
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'context': self.env.context,
        }

    def action_close(self):
        """Close wizard and return to quotation"""
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'sparks.solar.quotation',
            'res_id': self.quotation_id.id,
            'view_mode': 'form',
            'target': 'current',
        }
